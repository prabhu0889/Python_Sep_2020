# Step 1: install openpyxl:- pip install openpyxl

# Read data from Excel
import openpyxl

# Load the Workbook(excel file)
wb = openpyxl.load_workbook(r"D:\testlead.xlsx")

# returns all the sheet names of wb
print(wb.sheetnames)
# return active sheet name of wb
print(wb.active)

# create the object for specific sheet
sh = wb['Sheet1']

# for row count
print("total rows: ", sh.max_row)

# for column count
print("total column: ", sh.max_column)

# # way1
# print(sh['A4'].value)
for row in sh['A1' : 'D4']:
    for cell in row:
        print(cell.value)

# # way2
# for row in range(1, sh.max_row+1):           # for rows
#     for column in range(1, sh.max_column+1): # for column
#         print(sh.cell(row, column).value)
